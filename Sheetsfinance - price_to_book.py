from __future__ import annotations

import argparse
import json
import math
import os
import re
import time
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


GOOGLE_CREDENTIALS_PATH = Path("0_ingestion") / "stock-ingestion-494417-17cbf0e7891b.json"
EXPORT_WORKBOOK_PATH = Path("0_needs_processing") / "tickerlist.xlsx"
PROGRESS_PATH = Path("0_ingestion") / "price_to_book_progress.json"
RUN_STATE_PATH = Path("0_ingestion") / "price_to_book_last_run.json"

DEFAULT_SPREADSHEET_ID = "1dTYk1lIUlTp0cyvcYrCMgYywywuHCr6HNSDCZ87QyKw"
DEFAULT_SOURCE_WORKSHEET = "tickers"
DEFAULT_OUTPUT_WORKSHEET = "tickers_output"
DEFAULT_WORKSHEET_NAME = "Sheet1"
DEFAULT_POLL_SECONDS = 1.0
DEFAULT_TIMEOUT_SECONDS = 30.0
DEFAULT_MIN_WAIT_SECONDS = 1.0
DEFAULT_STABLE_READS = 1
DEFAULT_SPEEDRUN_WAIT_SECONDS = 240.0
DEFAULT_MAX_REFRESH_PASSES = 3
TICKER_HEADER = "ticker"
METRIC_HEADER = "price_to_book"

SHEET_ERROR_TOKENS = (
    "#ERROR!",
    "#VALUE!",
    "#DIV/0!",
    "#N/A",
    "#REF!",
    "#NAME?",
    "#NUM!",
    "#NULL!",
    "#FEIL!",
    "#VERDI!",
    "#I/T",
    "ERROR:",
    "OOPS",
    "LASTER INN",
    "LOADING",
    "FUNCTION TREND",
    "FUNCTION DIVIDE",
)


def excel_quote(value: str) -> str:
    return '"' + value.replace('"', '""') + '"'


def column_letter(column_number: int) -> str:
    if column_number <= 0:
        return ""
    letters: list[str] = []
    while column_number > 0:
        column_number, remainder = divmod(column_number - 1, 26)
        letters.append(chr(65 + remainder))
    return "".join(reversed(letters))


def normalize_header_name(value: Any) -> str:
    return str(value).strip().casefold()


def find_header_index(headers: list[str], required_header: str, worksheet_name: str) -> int:
    normalized_headers = [normalize_header_name(value) for value in headers]
    normalized_required = normalize_header_name(required_header)
    if normalized_required not in normalized_headers:
        raise ValueError(f"Required header {required_header!r} not found in worksheet {worksheet_name!r}.")
    return normalized_headers.index(normalized_required)


def is_bad_sheet_text(value: str) -> bool:
    upper_text = value.strip().upper()
    return any(token in upper_text for token in SHEET_ERROR_TOKENS)


def coerce_valid_number(value: Any) -> float | None:
    if value in ("", None):
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        number = float(value)
        return number if math.isfinite(number) else None
    if not isinstance(value, str):
        return None
    text = value.strip()
    if not text or is_bad_sheet_text(text):
        return None
    had_percent = "%" in text
    text = text.replace("\u2212", "-")
    text = text.replace("\xa0", " ")
    text = re.sub(r"[^0-9,\.\-+eE]", "", text)
    if not text:
        return None
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        number = float(text)
    except ValueError:
        return None
    if had_percent:
        number /= 100
    return number if math.isfinite(number) else None


def short_sheet_value(value: Any, max_length: int = 80) -> str:
    if value is None:
        return "blank"
    text = str(value).replace("\n", " ").strip()
    if not text:
        return "blank"
    if len(text) > max_length:
        return text[: max_length - 3] + "..."
    return text


def import_google_client():
    try:
        from google.oauth2.service_account import Credentials
        from googleapiclient.discovery import build
    except ModuleNotFoundError as exc:
        raise ModuleNotFoundError(
            "Google Sheets support requires 'google-auth' and 'google-api-python-client'."
        ) from exc
    return Credentials, build


def load_google_service(credentials_path: Path):
    if not credentials_path.exists():
        raise FileNotFoundError(f"Google service-account JSON not found: {credentials_path}")
    credentials_cls, build = import_google_client()
    scopes = ["https://www.googleapis.com/auth/spreadsheets"]
    credentials = credentials_cls.from_service_account_file(str(credentials_path), scopes=scopes)
    return build("sheets", "v4", credentials=credentials)


def get_google_header_cells(service, spreadsheet_id: str, sheet_name: str) -> dict[str, str]:
    response = service.spreadsheets().values().get(spreadsheetId=spreadsheet_id, range=f"{sheet_name}!1:1").execute()
    rows = response.get("values", [])
    if not rows:
        raise ValueError(f"Worksheet {sheet_name!r} does not contain a header row.")
    header_cells: dict[str, str] = {}
    for index, value in enumerate(rows[0], start=1):
        if value not in (None, ""):
            header_cells[str(value).strip()] = column_letter(index)
    return header_cells


def require_header_cell(header_cells: dict[str, str], required_header: str, worksheet_name: str) -> str:
    normalized_required = normalize_header_name(required_header)
    for header, column in header_cells.items():
        if normalize_header_name(header) == normalized_required:
            return column
    available_headers = ", ".join(repr(header) for header in header_cells) or "none"
    raise ValueError(
        f"Required header {required_header!r} not found in row 1 of worksheet {worksheet_name!r}. "
        f"Available headers: {available_headers}."
    )


def get_google_last_row(service, spreadsheet_id: str, sheet_name: str, ticker_column: str) -> int:
    response = service.spreadsheets().values().get(
        spreadsheetId=spreadsheet_id,
        range=f"{sheet_name}!{ticker_column}:{ticker_column}",
    ).execute()
    return len(response.get("values", []))


def get_google_values(service, spreadsheet_id: str, sheet_name: str) -> list[list[Any]]:
    response = service.spreadsheets().values().get(
        spreadsheetId=spreadsheet_id,
        range=sheet_name,
        valueRenderOption="UNFORMATTED_VALUE",
        dateTimeRenderOption="FORMATTED_STRING",
    ).execute()
    return response.get("values", [])


def get_google_range_values(service, spreadsheet_id: str, range_name: str, value_render_option: str = "UNFORMATTED_VALUE") -> list[list[Any]]:
    response = service.spreadsheets().values().get(
        spreadsheetId=spreadsheet_id,
        range=range_name,
        valueRenderOption=value_render_option,
        dateTimeRenderOption="FORMATTED_STRING",
    ).execute()
    return response.get("values", [])


def get_google_single_value(service, spreadsheet_id: str, range_name: str, value_render_option: str = "UNFORMATTED_VALUE") -> Any:
    values = get_google_range_values(service, spreadsheet_id, range_name, value_render_option=value_render_option)
    if values and values[0]:
        return values[0][0]
    return ""


def update_google_range_values(
    service,
    spreadsheet_id: str,
    range_name: str,
    values: list[list[str | int | float]],
    value_input_option: str = "RAW",
) -> None:
    service.spreadsheets().values().update(
        spreadsheetId=spreadsheet_id,
        range=range_name,
        valueInputOption=value_input_option,
        body={"values": values},
    ).execute()


def batch_update_google_ranges(
    service,
    spreadsheet_id: str,
    updates: list[tuple[str, list[list[str | int | float]]]],
    value_input_option: str = "RAW",
) -> None:
    if not updates:
        return
    service.spreadsheets().values().batchUpdate(
        spreadsheetId=spreadsheet_id,
        body={
            "valueInputOption": value_input_option,
            "data": [{"range": range_name, "values": values} for range_name, values in updates],
        },
    ).execute()


def normalize_rows(values: list[list[Any]]) -> list[list[Any]]:
    if not values:
        return []
    max_columns = max(len(row) for row in values)
    return [list(row) + [""] * (max_columns - len(row)) for row in values]


def merge_metric_column_to_excel(
    service,
    spreadsheet_id: str,
    output_sheet_name: str,
    export_path: Path,
    metric_header: str,
) -> int:
    values = normalize_rows(get_google_values(service, spreadsheet_id, output_sheet_name))
    if not values:
        raise ValueError(f"Worksheet {output_sheet_name!r} does not contain data.")

    output_headers = [str(value).strip() for value in values[0]]
    output_metric_idx = find_header_index(output_headers, metric_header, output_sheet_name)
    output_ticker_idx: int | None = None
    try:
        output_ticker_idx = find_header_index(output_headers, TICKER_HEADER, output_sheet_name)
    except ValueError:
        pass

    export_path.parent.mkdir(parents=True, exist_ok=True)
    if export_path.exists():
        workbook = load_workbook(export_path)
        ws = workbook[DEFAULT_WORKSHEET_NAME] if DEFAULT_WORKSHEET_NAME in workbook.sheetnames else workbook.active
    else:
        workbook = Workbook()
        ws = workbook.active
        ws.title = DEFAULT_WORKSHEET_NAME

    excel_headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
    occupied_header_width = 0
    for index, header in enumerate(excel_headers, start=1):
        if header:
            occupied_header_width = index
    normalized_excel_headers = [normalize_header_name(header) for header in excel_headers]
    normalized_metric_header = normalize_header_name(metric_header)
    if normalized_metric_header in normalized_excel_headers:
        metric_column = normalized_excel_headers.index(normalized_metric_header) + 1
    else:
        metric_column = occupied_header_width + 1
        ws.cell(row=1, column=metric_column, value=metric_header)

    excel_ticker_column: int | None = None
    normalized_ticker_header = normalize_header_name(TICKER_HEADER)
    if normalized_ticker_header in normalized_excel_headers:
        excel_ticker_column = normalized_excel_headers.index(normalized_ticker_header) + 1

    for row_number in range(2, ws.max_row + 1):
        ws.cell(row=row_number, column=metric_column, value=None)

    written_rows = 0
    if output_ticker_idx is not None and excel_ticker_column is not None:
        ticker_to_excel_row: dict[str, int] = {}
        for row_number in range(2, ws.max_row + 1):
            ticker = ws.cell(row=row_number, column=excel_ticker_column).value
            ticker_text = str(ticker).strip() if ticker is not None else ""
            if ticker_text:
                ticker_to_excel_row[ticker_text] = row_number

        for output_row in values[1:]:
            ticker = output_row[output_ticker_idx] if output_ticker_idx < len(output_row) else ""
            ticker_text = str(ticker).strip() if ticker is not None else ""
            if not ticker_text:
                continue
            value = output_row[output_metric_idx] if output_metric_idx < len(output_row) else ""
            target_row = ticker_to_excel_row.get(ticker_text)
            if target_row is None:
                target_row = ws.max_row + 1
                ws.cell(row=target_row, column=excel_ticker_column, value=ticker_text)
                ticker_to_excel_row[ticker_text] = target_row
            ws.cell(row=target_row, column=metric_column, value=value)
            written_rows += 1
    else:
        for offset, output_row in enumerate(values[1:], start=2):
            value = output_row[output_metric_idx] if output_metric_idx < len(output_row) else ""
            ws.cell(row=offset, column=metric_column, value=value)
            written_rows += 1

    workbook.save(export_path)
    return written_rows


def load_progress(progress_path: Path) -> dict[str, object]:
    if not progress_path.exists():
        return {}
    with progress_path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def save_progress(progress_path: Path, source_sheet_name: str, last_completed_row: int, last_completed_ticker: str) -> None:
    progress_path.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "source_worksheet": source_sheet_name,
        "last_completed_row": last_completed_row,
        "last_completed_ticker": last_completed_ticker,
        "updated_at": datetime.utcnow().isoformat(timespec="seconds") + "Z",
    }
    with progress_path.open("w", encoding="utf-8") as handle:
        json.dump(payload, handle, indent=2)


def clear_progress(progress_path: Path) -> None:
    if progress_path.exists():
        progress_path.unlink()


def load_run_state(run_state_path: Path) -> dict[str, object]:
    if not run_state_path.exists():
        return {}
    with run_state_path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def save_run_state(run_state_path: Path, spreadsheet_id: str, metric_header: str, rows_updated: int) -> None:
    run_state_path.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "spreadsheet_id": spreadsheet_id,
        "metric": metric_header,
        "run_date": datetime.now().date().isoformat(),
        "rows_updated": rows_updated,
        "updated_at": datetime.now().isoformat(timespec="seconds"),
    }
    with run_state_path.open("w", encoding="utf-8") as handle:
        json.dump(payload, handle, indent=2)


def already_ran_today(run_state_path: Path, spreadsheet_id: str, metric_header: str) -> bool:
    state = load_run_state(run_state_path)
    return (
        state.get("spreadsheet_id") == spreadsheet_id
        and state.get("metric") == metric_header
        and state.get("run_date") == datetime.now().date().isoformat()
    )


def ticker_ref(row: int, header_cells: dict[str, str]) -> str:
    ticker_column = require_header_cell(header_cells, TICKER_HEADER, DEFAULT_SOURCE_WORKSHEET)
    return f"{ticker_column}{row}"


def metric_formula(row: int, header_cells: dict[str, str]) -> str:
    return f'=SF({ticker_ref(row, header_cells)},{excel_quote("ratios")},{excel_quote("priceToBookRatio")},{excel_quote("ttm")})'


def wait_for_numeric_output(
    service,
    spreadsheet_id: str,
    range_name: str,
    poll_seconds: float,
    timeout_seconds: float,
    min_wait_seconds: float,
    stable_reads_required: int,
) -> tuple[float | None, Any]:
    start_time = time.time()
    last_number: float | None = None
    stable_reads = 0
    latest_raw: Any = ""
    stable_reads_required = max(1, stable_reads_required)
    poll_seconds = max(0.5, poll_seconds)
    while time.time() - start_time < timeout_seconds:
        elapsed = time.time() - start_time
        try:
            raw_output = get_google_single_value(service, spreadsheet_id, range_name, value_render_option="UNFORMATTED_VALUE")
        except Exception as exc:
            raw_output = f"Google read error: {exc}"
        latest_raw = raw_output
        number = coerce_valid_number(raw_output)
        if elapsed >= min_wait_seconds and number is not None:
            if last_number is not None and abs(number - last_number) < 1e-12:
                stable_reads += 1
            else:
                last_number = number
                stable_reads = 1
            if stable_reads >= stable_reads_required:
                return number, raw_output
        elif number is None:
            last_number = None
            stable_reads = 0
        time.sleep(poll_seconds)
    return None, latest_raw


def sync_metric_output_column(service, spreadsheet_id: str, source_sheet_name: str, output_sheet_name: str, header_name: str) -> int:
    source_values = get_google_values(service, spreadsheet_id, source_sheet_name)
    if not source_values:
        raise ValueError(f"Worksheet {source_sheet_name!r} does not contain data.")
    source_headers = [str(value).strip() for value in source_values[0]]
    source_idx = find_header_index(source_headers, header_name, source_sheet_name)
    output_header_cells = get_google_header_cells(service, spreadsheet_id, output_sheet_name)
    output_column = require_header_cell(output_header_cells, header_name, output_sheet_name)
    values: list[list[str | float]] = []
    copied_rows = 0
    for row in source_values[1:]:
        padded_row = list(row) + [""] * max(0, len(source_headers) - len(row))
        raw_value = padded_row[source_idx]
        number = coerce_valid_number(raw_value)
        values.append([raw_value if number is not None else ""])
        copied_rows += 1
    if values:
        update_google_range_values(service, spreadsheet_id, f"{output_sheet_name}!{output_column}2:{output_column}{copied_rows + 1}", values)
    return copied_rows


def run_metric_refresh(
    service,
    spreadsheet_id: str,
    source_sheet_name: str,
    output_sheet_name: str,
    poll_seconds: float,
    timeout_seconds: float,
    min_wait_seconds: float,
    stable_reads_required: int,
    speedrun_wait_seconds: float,
    retry_failures: bool,
    max_refresh_passes: int,
    progress_path: Path,
    resume_progress: bool,
) -> tuple[int, list[str]]:
    header_cells = get_google_header_cells(service, spreadsheet_id, source_sheet_name)
    output_header_cells = get_google_header_cells(service, spreadsheet_id, output_sheet_name)

    ticker_column = require_header_cell(header_cells, TICKER_HEADER, source_sheet_name)
    metric_column = require_header_cell(header_cells, METRIC_HEADER, source_sheet_name)
    output_column = require_header_cell(output_header_cells, METRIC_HEADER, output_sheet_name)
    last_row = get_google_last_row(service, spreadsheet_id, source_sheet_name, ticker_column)
    print(f"Starting {METRIC_HEADER}: worksheet={source_sheet_name}, rows 2-{last_row}, output={output_sheet_name}", flush=True)

    clear_progress(progress_path)
    resume_row = 2

    source_values = get_google_values(service, spreadsheet_id, source_sheet_name)
    source_headers = [str(value).strip() for value in source_values[0]]
    ticker_idx = find_header_index(source_headers, TICKER_HEADER, source_sheet_name)

    rows_to_process: list[tuple[int, str]] = []
    for row_number, row in enumerate(source_values[1:], start=2):
        if row_number < resume_row:
            continue
        padded_row = list(row) + [""] * max(0, len(source_headers) - len(row))
        ticker_text = str(padded_row[ticker_idx]).strip()
        if not ticker_text:
            break
        rows_to_process.append((row_number, ticker_text))

    pending_rows: list[tuple[int, str, Any]] = [(row_number, ticker_text, "") for row_number, ticker_text in rows_to_process]
    refreshed_rows = 0
    failures: list[str] = []
    max_refresh_passes = max(1, max_refresh_passes)
    if not retry_failures:
        max_refresh_passes = 1

    for pass_number in range(1, max_refresh_passes + 1):
        if not pending_rows:
            break

        formula_updates = [
            (f"{source_sheet_name}!{metric_column}{row_number}", [[metric_formula(row_number, header_cells)]])
            for row_number, _, _ in pending_rows
        ]
        if pass_number == 1:
            print(f"Refresh pass {pass_number}/{max_refresh_passes}: refreshing {len(formula_updates)} ticker cells.", flush=True)
        else:
            print(f"Refresh pass {pass_number}/{max_refresh_passes}: refreshing {len(formula_updates)} error/not-ready cells.", flush=True)
        batch_update_google_ranges(service, spreadsheet_id, formula_updates, value_input_option="USER_ENTERED")

        if speedrun_wait_seconds > 0:
            print(f"Refresh pass {pass_number}/{max_refresh_passes}: waiting {speedrun_wait_seconds:g}s before collecting values.", flush=True)
            time.sleep(speedrun_wait_seconds)

        read_start_row = pending_rows[0][0]
        read_end_row = pending_rows[-1][0]
        pass_values = get_google_range_values(
            service,
            spreadsheet_id,
            f"{source_sheet_name}!{metric_column}{read_start_row}:{metric_column}{read_end_row}",
            value_render_option="UNFORMATTED_VALUE",
        )

        output_updates: list[tuple[str, list[list[str | int | float]]]] = []
        next_pending_rows: list[tuple[int, str, Any]] = []
        for row_number, ticker_text, previous_raw_value in pending_rows:
            retry_value_index = row_number - read_start_row
            raw_value = (
                pass_values[retry_value_index][0]
                if retry_value_index < len(pass_values) and pass_values[retry_value_index]
                else ""
            )
            number = coerce_valid_number(raw_value)
            if number is None:
                next_pending_rows.append((row_number, ticker_text, raw_value or previous_raw_value))
                continue
            output_updates.append((f"{output_sheet_name}!{output_column}{row_number}", [[number]]))
            refreshed_rows += 1
            save_progress(progress_path, source_sheet_name, row_number, ticker_text)

        print(
            f"Refresh pass {pass_number}/{max_refresh_passes}: captured {len(output_updates)} values, {len(next_pending_rows)} still not ready/error.",
            flush=True,
        )
        batch_update_google_ranges(service, spreadsheet_id, output_updates)
        pending_rows = next_pending_rows

    failures = [f"{ticker}: {short_sheet_value(raw_value)}" for _, ticker, raw_value in pending_rows]

    clear_progress(progress_path)
    return refreshed_rows, failures


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Refresh price_to_book in tickers cell-by-cell, sync it to tickers_output, and export tickerlist.xlsx.")
    parser.add_argument("--sheet-id", default=os.getenv("GOOGLE_SHEET_ID", DEFAULT_SPREADSHEET_ID))
    parser.add_argument("--credentials", default=str(GOOGLE_CREDENTIALS_PATH))
    parser.add_argument("--source-worksheet", default=os.getenv("INGESTION_SOURCE_WORKSHEET", DEFAULT_SOURCE_WORKSHEET))
    parser.add_argument("--output-worksheet", default=os.getenv("INGESTION_OUTPUT_WORKSHEET", DEFAULT_OUTPUT_WORKSHEET))
    parser.add_argument("--export-path", default=str(EXPORT_WORKBOOK_PATH))
    parser.add_argument("--poll-seconds", type=float, default=float(os.getenv("INGESTION_PRICE_TO_BOOK_POLL_SECONDS", str(DEFAULT_POLL_SECONDS))))
    parser.add_argument("--timeout-seconds", type=float, default=float(os.getenv("INGESTION_PRICE_TO_BOOK_TIMEOUT_SECONDS", str(DEFAULT_TIMEOUT_SECONDS))))
    parser.add_argument("--min-wait-seconds", type=float, default=float(os.getenv("INGESTION_PRICE_TO_BOOK_MIN_WAIT_SECONDS", str(DEFAULT_MIN_WAIT_SECONDS))))
    parser.add_argument("--stable-reads", type=int, default=int(os.getenv("INGESTION_PRICE_TO_BOOK_STABLE_READS", str(DEFAULT_STABLE_READS))))
    parser.add_argument("--speedrun-wait-seconds", type=float, default=float(os.getenv("INGESTION_PRICE_TO_BOOK_SPEEDRUN_WAIT_SECONDS", str(DEFAULT_SPEEDRUN_WAIT_SECONDS))))
    parser.add_argument("--max-refresh-passes", type=int, default=int(os.getenv("INGESTION_PRICE_TO_BOOK_MAX_REFRESH_PASSES", str(DEFAULT_MAX_REFRESH_PASSES))))
    parser.add_argument("--no-retry-failures", action="store_true")
    parser.add_argument("--progress-path", default=str(PROGRESS_PATH))
    parser.add_argument("--run-state-path", default=str(RUN_STATE_PATH))
    parser.add_argument("--force", action="store_true", help="Run even if this sheet/metric already completed today.")
    parser.add_argument("--no-resume-progress", action="store_true")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if not args.force and already_ran_today(Path(args.run_state_path), args.sheet_id, METRIC_HEADER):
        print(
            f"Skipping {METRIC_HEADER}: sheet {args.sheet_id} already completed today. Use --force to rerun.",
            flush=True,
        )
        return

    print(f"Using Google Sheet ID: {args.sheet_id}", flush=True)
    print(f"Loading Google credentials: {args.credentials}", flush=True)
    service = load_google_service(Path(args.credentials))
    print("Google Sheets service loaded.", flush=True)
    print(f"Reading worksheets: {args.source_worksheet}, {args.output_worksheet}", flush=True)
    refreshed_rows, failures = run_metric_refresh(
        service=service,
        spreadsheet_id=args.sheet_id,
        source_sheet_name=args.source_worksheet,
        output_sheet_name=args.output_worksheet,
        poll_seconds=args.poll_seconds,
        timeout_seconds=args.timeout_seconds,
        min_wait_seconds=args.min_wait_seconds,
        stable_reads_required=args.stable_reads,
        speedrun_wait_seconds=args.speedrun_wait_seconds,
        retry_failures=not args.no_retry_failures,
        max_refresh_passes=args.max_refresh_passes,
        progress_path=Path(args.progress_path),
        resume_progress=not args.no_resume_progress,
    )
    copied_rows = refreshed_rows
    exported_rows = merge_metric_column_to_excel(
        service,
        args.sheet_id,
        args.output_worksheet,
        Path(args.export_path),
        METRIC_HEADER,
    )

    print(f"Metric refreshed: {METRIC_HEADER}")
    print(f"Rows refreshed in tickers: {refreshed_rows}")
    print(f"Rows written to tickers_output: {copied_rows}")
    print(f"Excel rows updated for {METRIC_HEADER}: {exported_rows}")
    print(f"Export path: {args.export_path}")
    save_run_state(Path(args.run_state_path), args.sheet_id, METRIC_HEADER, refreshed_rows)
    if failures:
        print(f"Rows that did not resolve to a valid {METRIC_HEADER} value:")
        for failure in failures:
            print(f"- {failure}")


if __name__ == "__main__":
    main()

