from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


SHEET_ID = "1O8VNU8ykEnlbgkuebKrv2u0V5vX5Ib4EBs2mLC3xZBU"
OUTPUT_WORKSHEET = "tickers_output"
GOOGLE_CREDENTIALS_PATH = Path("0_ingestion") / "stock-ingestion-494417-17cbf0e7891b.json"
EXPORT_WORKBOOK_PATH = Path("0_needs_processing") / "tickerlist.xlsx"
LOCAL_WORKSHEET_NAME = "Sheet1"
TICKER_HEADER = "ticker"
METRIC_HEADER = "3_mth_momentum"


def normalize_header(value: Any) -> str:
    return str(value).strip().casefold()


def load_google_service(credentials_path: Path):
    from google.oauth2.service_account import Credentials
    from googleapiclient.discovery import build

    credentials = Credentials.from_service_account_file(
        str(credentials_path),
        scopes=["https://www.googleapis.com/auth/spreadsheets.readonly"],
    )
    return build("sheets", "v4", credentials=credentials)


def read_sheet_values(service) -> list[list[Any]]:
    response = service.spreadsheets().values().get(
        spreadsheetId=SHEET_ID,
        range=f"'{OUTPUT_WORKSHEET}'!1:10000",
        valueRenderOption="UNFORMATTED_VALUE",
    ).execute()
    return response.get("values", [])


def header_index(headers: list[Any], header: str) -> int:
    normalized = [normalize_header(value) for value in headers]
    wanted = normalize_header(header)
    if wanted not in normalized:
        available = ", ".join(str(value) for value in headers)
        raise ValueError(f"Header {header!r} not found. Available headers: {available}")
    return normalized.index(wanted)


def ensure_workbook(path: Path):
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        workbook = load_workbook(path)
        worksheet = workbook[LOCAL_WORKSHEET_NAME] if LOCAL_WORKSHEET_NAME in workbook.sheetnames else workbook.active
    else:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = LOCAL_WORKSHEET_NAME
    return workbook, worksheet


def ensure_column(worksheet, header: str) -> int:
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in worksheet[1]]
    normalized = [normalize_header(value) for value in headers]
    wanted = normalize_header(header)
    if wanted in normalized:
        return normalized.index(wanted) + 1
    occupied = max((index for index, value in enumerate(headers, start=1) if value), default=0)
    column = occupied + 1
    worksheet.cell(row=1, column=column, value=header)
    return column


def ticker_row_map(worksheet) -> tuple[int, dict[str, int]]:
    ticker_column = ensure_column(worksheet, TICKER_HEADER)
    rows_by_ticker: dict[str, int] = {}
    for row_number in range(2, worksheet.max_row + 1):
        ticker = worksheet.cell(row=row_number, column=ticker_column).value
        ticker_text = str(ticker).strip().upper() if ticker is not None else ""
        if ticker_text:
            rows_by_ticker[ticker_text] = row_number
    return ticker_column, rows_by_ticker


def merge_momentum(worksheet, values: list[list[Any]]) -> int:
    if not values:
        raise ValueError("Output worksheet is empty.")

    headers = values[0]
    ticker_idx = header_index(headers, TICKER_HEADER)
    metric_idx = header_index(headers, METRIC_HEADER)
    source_nonempty = sum(
        1
        for row in values[1:]
        if metric_idx < len(row) and row[metric_idx] not in ("", None)
    )
    if source_nonempty == 0:
        raise ValueError(f"{OUTPUT_WORKSHEET!r} has no filled {METRIC_HEADER!r} values.")

    ticker_column, rows_by_ticker = ticker_row_map(worksheet)
    metric_column = ensure_column(worksheet, METRIC_HEADER)

    for row_number in range(2, worksheet.max_row + 1):
        worksheet.cell(row=row_number, column=metric_column, value=None)

    written = 0
    for row in values[1:]:
        ticker = row[ticker_idx] if ticker_idx < len(row) else ""
        ticker_text = str(ticker).strip().upper() if ticker is not None else ""
        if not ticker_text:
            continue

        value = row[metric_idx] if metric_idx < len(row) else ""
        target_row = rows_by_ticker.get(ticker_text)
        if target_row is None:
            target_row = worksheet.max_row + 1
            worksheet.cell(row=target_row, column=ticker_column, value=ticker_text)
            rows_by_ticker[ticker_text] = target_row

        worksheet.cell(row=target_row, column=metric_column, value=value)
        written += 1

    return written


def main() -> None:
    service = load_google_service(GOOGLE_CREDENTIALS_PATH)
    values = read_sheet_values(service)
    workbook, worksheet = ensure_workbook(EXPORT_WORKBOOK_PATH)
    written = merge_momentum(worksheet, values)
    workbook.save(EXPORT_WORKBOOK_PATH)
    print(f"Merged {written} rows into {EXPORT_WORKBOOK_PATH} column {METRIC_HEADER!r}.")


if __name__ == "__main__":
    main()
