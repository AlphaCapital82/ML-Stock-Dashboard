from __future__ import annotations

import argparse
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


GOOGLE_CREDENTIALS_PATH = Path("0_ingestion") / "stock-ingestion-494417-17cbf0e7891b.json"
EXPORT_WORKBOOK_PATH = Path("0_needs_processing") / "tickerlist.xlsx"
DEFAULT_OUTPUT_WORKSHEET = "tickers_output"
DEFAULT_WORKSHEET_NAME = "Sheet1"
TICKER_HEADER = "ticker"

METRIC_SHEETS = [
    {
        "name": "3_mth_momentum",
        "sheet_id": "1O8VNU8ykEnlbgkuebKrv2u0V5vX5Ib4EBs2mLC3xZBU",
        "metric_header": "3_mth_momentum",
    },
    {
        "name": "debt_to_assets",
        "sheet_id": "1z_j9QHFcw5m47GQsduOIwEabPitRHkGcM7u_aQGmz1g",
        "metric_header": "debt_to_assets",
    },
    {
        "name": "earnings_surprise",
        "sheet_id": "1PHYPBjp-sbLYiep5K4b9tTIRj877mwfPIRcQvc82M3c",
        "metric_header": "earnings_surprise",
    },
    {
        "name": "ev_to_ebitda",
        "sheet_id": "1mfIcxXzMmJuD1ljsOOtkOO74qoChS-SMHbO-2561hpU",
        "metric_header": "ev_to_ebitda",
    },
    {
        "name": "price_to_book",
        "sheet_id": "1dTYk1lIUlTp0cyvcYrCMgYywywuHCr6HNSDCZ87QyKw",
        "metric_header": "price_to_book",
        "output_worksheet": "Sheet1",
    },
    {
        "name": "price_to_earnings",
        "sheet_id": "1h6hyiBGSJvXNLAh_7PTTo_4Qc3f_KJg8JuFXT2zY1Gk",
        "metric_header": "price_to_earnings",
        "output_worksheet": "Sheet1",
    },
    {
        "name": "roic",
        "sheet_id": "1jQNqmE8X33OMfEWlqrLZ-rwRUIrg_4CxU-Mfn5vRqbM",
        "metric_header": "ROIC",
        "local_header": "roic",
    },
    {
        "name": "stock_volatility",
        "sheet_id": "11G8OO8_XhHCh1-nWY7rGTMMoPjTb0svMiYCxBQb9YWk",
        "metric_header": "stock_volatility",
    },
    {
        "name": "valuation_gap",
        "sheet_id": "1FBnRzytDx-5uNRmK4Qqagt8CsnKRCkiI2WUiyZMfOIs",
        "metric_header": "valuation_gap",
    },
]


def normalize_header(value: Any) -> str:
    return str(value).strip().casefold()


def import_google_client():
    from google.oauth2.service_account import Credentials
    from googleapiclient.discovery import build

    return Credentials, build


def load_google_service(credentials_path: Path):
    credentials_cls, build = import_google_client()
    credentials = credentials_cls.from_service_account_file(
        str(credentials_path),
        scopes=["https://www.googleapis.com/auth/spreadsheets.readonly"],
    )
    return build("sheets", "v4", credentials=credentials)


def read_output_values(service, sheet_id: str, output_worksheet: str) -> list[list[Any]]:
    response = service.spreadsheets().values().get(
        spreadsheetId=sheet_id,
        range=f"'{output_worksheet}'!1:10000",
        valueRenderOption="UNFORMATTED_VALUE",
    ).execute()
    return response.get("values", [])


def header_index(headers: list[Any], header: str) -> int:
    normalized = [normalize_header(value) for value in headers]
    wanted = normalize_header(header)
    if wanted not in normalized:
        available = ", ".join(str(value) for value in headers)
        raise ValueError(f"Header {header!r} not found. Available headers: {available}")
    return normalized.index(wanted)


def ensure_workbook(path: Path):
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        workbook = load_workbook(path)
        worksheet = workbook[DEFAULT_WORKSHEET_NAME] if DEFAULT_WORKSHEET_NAME in workbook.sheetnames else workbook.active
    else:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = DEFAULT_WORKSHEET_NAME
    return workbook, worksheet


def ensure_column(worksheet, header: str) -> int:
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in worksheet[1]]
    normalized = [normalize_header(value) for value in headers]
    wanted = normalize_header(header)
    if wanted in normalized:
        return normalized.index(wanted) + 1
    occupied = max((index for index, value in enumerate(headers, start=1) if value), default=0)
    column = occupied + 1
    worksheet.cell(row=1, column=column, value=header)
    return column


def ticker_row_map(worksheet) -> tuple[int, dict[str, int]]:
    ticker_column = ensure_column(worksheet, TICKER_HEADER)
    mapping: dict[str, int] = {}
    for row_number in range(2, worksheet.max_row + 1):
        ticker = worksheet.cell(row=row_number, column=ticker_column).value
        ticker_text = str(ticker).strip().upper() if ticker is not None else ""
        if ticker_text:
            mapping[ticker_text] = row_number
    return ticker_column, mapping


def merge_metric(worksheet, values: list[list[Any]], metric_header: str, local_header: str) -> int:
    if not values:
        raise ValueError("Output worksheet is empty.")
    headers = values[0]
    ticker_idx = header_index(headers, TICKER_HEADER)
    metric_idx = header_index(headers, metric_header)
    source_nonempty = sum(
        1
        for row in values[1:]
        if metric_idx < len(row) and row[metric_idx] not in ("", None)
    )
    if source_nonempty == 0:
        raise ValueError(f"Output worksheet has no filled {metric_header!r} values.")

    ticker_column, rows_by_ticker = ticker_row_map(worksheet)
    metric_column = ensure_column(worksheet, local_header)

    for row_number in range(2, worksheet.max_row + 1):
        worksheet.cell(row=row_number, column=metric_column, value=None)

    written = 0
    for row in values[1:]:
        ticker = row[ticker_idx] if ticker_idx < len(row) else ""
        ticker_text = str(ticker).strip().upper() if ticker is not None else ""
        if not ticker_text:
            continue
        value = row[metric_idx] if metric_idx < len(row) else ""
        target_row = rows_by_ticker.get(ticker_text)
        if target_row is None:
            target_row = worksheet.max_row + 1
            worksheet.cell(row=target_row, column=ticker_column, value=ticker_text)
            rows_by_ticker[ticker_text] = target_row
        worksheet.cell(row=target_row, column=metric_column, value=value)
        written += 1
    return written


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Merge metric-specific SheetsFinance tickers_output tabs into local tickerlist.xlsx without refreshing formulas."
    )
    parser.add_argument("--credentials", default=str(GOOGLE_CREDENTIALS_PATH))
    parser.add_argument("--export-path", default=str(EXPORT_WORKBOOK_PATH))
    parser.add_argument("--output-worksheet", default=DEFAULT_OUTPUT_WORKSHEET)
    parser.add_argument("--metric", action="append", help="Metric name to sync. Can be repeated. Defaults to all configured metrics.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    selected = set(args.metric or [])
    metrics = [metric for metric in METRIC_SHEETS if not selected or metric["name"] in selected]
    if not metrics:
        configured = ", ".join(metric["name"] for metric in METRIC_SHEETS)
        raise ValueError(f"No matching metrics selected. Configured metrics: {configured}")

    service = load_google_service(Path(args.credentials))
    workbook, worksheet = ensure_workbook(Path(args.export_path))

    for metric in metrics:
        worksheet_name = metric.get("output_worksheet", args.output_worksheet)
        try:
            values = read_output_values(service, metric["sheet_id"], worksheet_name)
            written = merge_metric(
                worksheet,
                values,
                metric["metric_header"],
                metric.get("local_header", metric["metric_header"]),
            )
            workbook.save(args.export_path)
            print(f"{metric['name']}: merged {written} rows from {worksheet_name}")
        except Exception as exc:
            workbook.save(args.export_path)
            print(f"{metric['name']}: skipped ({exc})")

    print(f"Saved {args.export_path}")


if __name__ == "__main__":
    main()
