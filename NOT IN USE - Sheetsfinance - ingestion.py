from __future__ import annotations

import argparse
import json
import math
import os
import re
import time
from datetime import datetime
from pathlib import Path
from typing import Callable, Any

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet


WORKBOOK_PATH = Path("0_ingestion") / "tickerlist.xlsx"
GOOGLE_CREDENTIALS_PATH = Path("0_ingestion") / "stock-ingestion-494417-17cbf0e7891b.json"
EXPORT_WORKBOOK_PATH = Path("0_needs_processing") / "tickerlist.xlsx"
PROGRESS_PATH = Path("0_ingestion") / "valuation_gap_progress.json"

DEFAULT_SPREADSHEET_ID = "1FBnRzytDx-5uNRmK4Qqagt8CsnKRCkiI2WUiyZMfOIs"
DEFAULT_WORKSHEET_NAME = "Sheet1"
DEFAULT_SOURCE_WORKSHEET = "tickers"
DEFAULT_OUTPUT_WORKSHEET = "tickers_output"
DEFAULT_MAX_OUTPUT_COLUMNS = 13
DEFAULT_DCF_WORKSHEET = "dcf"
DEFAULT_VALUATION_GAP_HEADER = "valuation_gap"
DEFAULT_DCF_INPUT_CELL = "A2"
DEFAULT_DCF_OUTPUT_CELL = "B21"
DEFAULT_DCF_POLL_SECONDS = 4.0
DEFAULT_DCF_TIMEOUT_SECONDS = 90.0
DEFAULT_DCF_MIN_WAIT_SECONDS = 8.0
DEFAULT_DCF_STABLE_READS = 2

TICKER_HEADER = "Ticker"

# Explicit year choices keep the formulas readable and aligned to the workbook's headers.
REVENUE_GROWTH_YEAR = 2024
ESTIMATE_GROWTH_YEAR = 2025
LAST_FULL_CALENDAR_YEAR = datetime.now().year - 1


FormulaBuilder = Callable[[int, dict[str, str], int], str]


SHEET_ERROR_TOKENS = (
    "#ERROR!",
    "#VALUE!",
    "#DIV/0!",
    "#N/A",
    "#REF!",
    "#NAME?",
    "#NUM!",
    "#NULL!",
    "#FEIL!",
    "#VERDI!",
    "#I/T",
    "ERROR:",
    "OOPS",
    "LASTER INN",
    "LOADING",
    "FUNCTION TREND",
    "FUNCTION DIVIDE",
)


def excel_quote(value: str) -> str:
    return '"' + value.replace('"', '""') + '"'


def column_letter(column_number: int) -> str:
    if column_number <= 0:
        return ""

    letters: list[str] = []
    while column_number > 0:
        column_number, remainder = divmod(column_number - 1, 26)
        letters.append(chr(65 + remainder))
    return "".join(reversed(letters))


def ticker_ref(row: int, header_cells: dict[str, str]) -> str:
    return f"${header_cells[TICKER_HEADER]}{row}"


def sf_formula(row: int, header_cells: dict[str, str], *parts: str) -> str:
    args = [ticker_ref(row, header_cells)]
    args.extend(excel_quote(part) for part in parts)
    return "=SF(" + ",".join(args) + ")"


def sf_timeseries_formula(row: int, header_cells: dict[str, str], metric: str) -> str:
    return (
        f'=SF_TIMESERIES({ticker_ref(row, header_cells)},TODAY()-90,TODAY(),"",'
        f'{excel_quote(metric)},{excel_quote("NH")})'
    )


def index_formula(base_formula: str) -> str:
    return f"INDEX({base_formula},1,1)"


def wrap_iferror(formula: str, fallback: str = '""') -> str:
    return f"=IFERROR({formula},{fallback})"


def direct_sf(*parts: str) -> FormulaBuilder:
    def builder(row: int, header_cells: dict[str, str], _: int) -> str:
        return wrap_iferror(sf_formula(row, header_cells, *parts)[1:])

    return builder


def earnings_surprise_formula(row: int, header_cells: dict[str, str], _: int) -> str:
    actual = index_formula(sf_formula(row, header_cells, "earnings", "eps", "ttm", "NH")[1:])
    estimate = index_formula(sf_formula(row, header_cells, "earnings", "epsEstimated", "ttm", "NH")[1:])
    return wrap_iferror(f"{actual}-{estimate}")


def estimate_growth_formula(row: int, header_cells: dict[str, str], _: int) -> str:
    estimate_2025 = sf_formula(row, header_cells, "estimates", "revenueAvg", str(ESTIMATE_GROWTH_YEAR))
    revenue_2024 = sf_formula(row, header_cells, "income", "revenue", str(REVENUE_GROWTH_YEAR), "calYear")
    return wrap_iferror(f"({estimate_2025[1:]}/{revenue_2024[1:]})-1")


def valuation_gap_formula(row: int, header_cells: dict[str, str], _: int) -> str:
    latest_target = index_formula(sf_formula(row, header_cells, "priceTargets", "priceTarget", "12", "NH")[1:])
    current_price = sf_formula(row, header_cells, "realTime", "price")[1:]
    return wrap_iferror(f"({latest_target}/{current_price})-1")


def gross_profit_formula(quarter_suffix: str) -> FormulaBuilder:
    def builder(row: int, header_cells: dict[str, str], _: int) -> str:
        return wrap_iferror(
            sf_formula(
                row,
                header_cells,
                f"income{quarter_suffix}",
                "grossProfit",
                str(LAST_FULL_CALENDAR_YEAR),
                "calYear",
            )[1:]
        )

    return builder


def quarterly_growth_formula(row: int, header_cells: dict[str, str], _: int) -> str:
    q2_ref = f"${header_cells['Q2 gross profit']}{row}"
    q3_ref = f"${header_cells['Q3 gross profit']}{row}"
    return wrap_iferror(f"({q3_ref}/{q2_ref})-1")


def price_change_formula(metric: str) -> FormulaBuilder:
    def builder(row: int, header_cells: dict[str, str], _: int) -> str:
        return wrap_iferror(sf_formula(row, header_cells, "change", metric, "", "decimal")[1:])

    return builder


def stock_volatility_formula(row: int, header_cells: dict[str, str], _: int) -> str:
    series = sf_timeseries_formula(row, header_cells, "changePercent")[1:]
    return wrap_iferror(f"STDEV.P({series})/100")


def rank_formula(row: int, header_cells: dict[str, str], last_row: int) -> str:
    if "valuation_gap" not in header_cells:
        return ""
    valuation_col = header_cells["valuation_gap"]
    current = f"${valuation_col}{row}"
    full_range = f"${valuation_col}$2:${valuation_col}${last_row}"
    return wrap_iferror(f"RANK.EQ({current},{full_range},0)")


FORMULA_BUILDERS: dict[str, FormulaBuilder] = {
    "Name": direct_sf("companyInfo", "name"),
    "Country": direct_sf("companyInfo", "country"),
    "Industry": direct_sf("companyInfo", "industry"),
    "Sector": direct_sf("companyInfo", "sector"),
    "ROIC": direct_sf("ratios", "returnOnInvestedCapital", "ttm"),
    "rev_growth_2024": direct_sf("growth", "revenueGrowth", str(REVENUE_GROWTH_YEAR)),
    "debt_to_assets": direct_sf("ratios", "debtToAssetsRatio", "ttm"),
    "ev_to_ebitda": direct_sf("ratios", "evToEBITDA", "ttm"),
    "price_to_book": direct_sf("ratios", "priceToBookRatio", "ttm"),
    "Earnings surprise": earnings_surprise_formula,
    "API_GROWTH_2025": estimate_growth_formula,
    "valuation_gap": valuation_gap_formula,
    "Q2 gross profit": gross_profit_formula("Q2"),
    "Q3 gross profit": gross_profit_formula("Q3"),
    "Quarterly growth": quarterly_growth_formula,
    "1 month": price_change_formula("1M"),
    "3_month": price_change_formula("3M"),
    "12_month": price_change_formula("1Y"),
    "Rank": rank_formula,
    "stock_volatility": stock_volatility_formula,
}


def normalize_header_name(value: Any) -> str:
    return str(value).strip().casefold()


def find_header_index(headers: list[str], required_header: str, worksheet_name: str) -> int:
    normalized_headers = [normalize_header_name(value) for value in headers]
    normalized_required = normalize_header_name(required_header)
    if normalized_required not in normalized_headers:
        raise ValueError(f"Required header {required_header!r} not found in worksheet {worksheet_name!r}.")
    return normalized_headers.index(normalized_required)


def is_bad_sheet_text(value: str) -> bool:
    upper_text = value.strip().upper()
    return any(token in upper_text for token in SHEET_ERROR_TOKENS)


def coerce_valid_number(value: Any) -> float | None:
    """Return a finite float only when Sheets returned a clean numeric result.

    This deliberately rejects Google Sheets / SheetsFinance error states such as
    #DIV/0!, #VALUE!, #ERROR!, "Laster inn ...", and "Loading data".
    """
    if value in ("", None):
        return None

    if isinstance(value, bool):
        return None

    if isinstance(value, (int, float)):
        number = float(value)
        return number if math.isfinite(number) else None

    if not isinstance(value, str):
        return None

    text = value.strip()
    if not text or is_bad_sheet_text(text):
        return None

    had_percent = "%" in text

    # Keep only numeric characters/signs/decimal separators/exponent notation.
    text = text.replace("\u2212", "-")  # Unicode minus
    text = text.replace("\xa0", " ")
    text = re.sub(r"[^0-9,\.\-+eE]", "", text)

    if not text:
        return None

    # Support both Norwegian and English number formatting:
    # 1,23 -> 1.23
    # 1.23 -> 1.23
    # 1 234,56 / 1.234,56 -> 1234.56
    # 1,234.56 -> 1234.56
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            # Decimal comma, dot thousands.
            text = text.replace(".", "").replace(",", ".")
        else:
            # Decimal dot, comma thousands.
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(",", ".")

    try:
        number = float(text)
    except ValueError:
        return None

    if had_percent:
        number = number / 100

    return number if math.isfinite(number) else None


def short_sheet_value(value: Any, max_length: int = 80) -> str:
    if value is None:
        return "blank"
    text = str(value).replace("\n", " ").strip()
    if not text:
        return "blank"
    if len(text) > max_length:
        return text[: max_length - 3] + "..."
    return text


def get_sheet(sheet_name: str) -> Worksheet:
    workbook = load_workbook(WORKBOOK_PATH)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Worksheet {sheet_name!r} not found in {WORKBOOK_PATH}")
    return workbook[sheet_name]


def read_header_cells(ws: Worksheet) -> dict[str, str]:
    header_cells: dict[str, str] = {}
    for cell in ws[1]:
        if cell.value is None:
            continue
        header_cells[str(cell.value).strip()] = cell.column_letter
    return header_cells


def populate_formulas(ws: Worksheet) -> tuple[int, list[str], list[str]]:
    header_cells = read_header_cells(ws)
    if TICKER_HEADER not in header_cells:
        raise ValueError(f"Required header {TICKER_HEADER!r} not found in row 1.")

    missing_supported_headers = [header for header in FORMULA_BUILDERS if header not in header_cells]
    unmapped_headers = [header for header in header_cells if header not in FORMULA_BUILDERS and header != TICKER_HEADER]
    last_row = ws.max_row

    for row in range(2, last_row + 1):
        ticker_cell = ws[f"{header_cells[TICKER_HEADER]}{row}"]
        if ticker_cell.value in (None, ""):
            continue

        for header, builder in FORMULA_BUILDERS.items():
            column = header_cells.get(header)
            if not column:
                continue

            formula = builder(row, header_cells, last_row)
            if not formula:
                continue
            ws[f"{column}{row}"] = formula

    return last_row - 1, missing_supported_headers, unmapped_headers


def autosize_columns(ws: Worksheet) -> None:
    for column_cells in ws.columns:
        letter = column_cells[0].column_letter
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells[:20])
        ws.column_dimensions[letter].width = min(max(max_length + 2, 10), 28)


def import_google_client():
    try:
        from google.oauth2.service_account import Credentials
        from googleapiclient.discovery import build
    except ModuleNotFoundError as exc:
        raise ModuleNotFoundError(
            "Google Sheets support requires 'google-auth' and 'google-api-python-client'. "
            "Install them in the active environment before using --backend gsheet."
        ) from exc
    return Credentials, build


def load_google_service(credentials_path: Path):
    if not credentials_path.exists():
        raise FileNotFoundError(f"Google service-account JSON not found: {credentials_path}")

    Credentials, build = import_google_client()
    scopes = ["https://www.googleapis.com/auth/spreadsheets"]
    credentials = Credentials.from_service_account_file(str(credentials_path), scopes=scopes)
    return build("sheets", "v4", credentials=credentials)


def get_sheet_metadata(service, spreadsheet_id: str) -> list[dict]:
    response = service.spreadsheets().get(spreadsheetId=spreadsheet_id).execute()
    return response.get("sheets", [])


def get_google_sheet_id(sheet_metadata: list[dict], sheet_name: str) -> int:
    for sheet in sheet_metadata:
        properties = sheet.get("properties", {})
        if properties.get("title") == sheet_name:
            return int(properties["sheetId"])
    raise ValueError(f"Worksheet {sheet_name!r} not found in Google Sheet.")


def get_google_header_cells(service, spreadsheet_id: str, sheet_name: str) -> dict[str, str]:
    response = (
        service.spreadsheets()
        .values()
        .get(spreadsheetId=spreadsheet_id, range=f"{sheet_name}!1:1")
        .execute()
    )
    rows = response.get("values", [])
    if not rows:
        raise ValueError(f"Worksheet {sheet_name!r} does not contain a header row.")

    header_cells: dict[str, str] = {}
    for index, value in enumerate(rows[0], start=1):
        if value in (None, ""):
            continue
        header_cells[str(value).strip()] = column_letter(index)
    return header_cells


def get_google_last_row(service, spreadsheet_id: str, sheet_name: str, ticker_column: str) -> int:
    response = (
        service.spreadsheets()
        .values()
        .get(spreadsheetId=spreadsheet_id, range=f"{sheet_name}!{ticker_column}:{ticker_column}")
        .execute()
    )
    values = response.get("values", [])
    return len(values)


def get_google_values(service, spreadsheet_id: str, sheet_name: str) -> list[list[Any]]:
    response = (
        service.spreadsheets()
        .values()
        .get(
            spreadsheetId=spreadsheet_id,
            range=sheet_name,
            valueRenderOption="UNFORMATTED_VALUE",
            dateTimeRenderOption="FORMATTED_STRING",
        )
        .execute()
    )
    return response.get("values", [])


def normalize_rows(values: list[list[Any]]) -> list[list[Any]]:
    if not values:
        return []

    max_columns = max(len(row) for row in values)
    normalized: list[list[Any]] = []
    for row in values:
        padded = list(row) + [""] * (max_columns - len(row))
        normalized.append(padded)
    return normalized


def load_progress(progress_path: Path) -> dict[str, object]:
    if not progress_path.exists():
        return {}
    with progress_path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def save_progress(
    progress_path: Path,
    source_sheet_name: str,
    last_completed_row: int,
    last_completed_ticker: str,
) -> None:
    progress_path.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "source_worksheet": source_sheet_name,
        "last_completed_row": last_completed_row,
        "last_completed_ticker": last_completed_ticker,
        "updated_at": datetime.utcnow().isoformat(timespec="seconds") + "Z",
    }
    with progress_path.open("w", encoding="utf-8") as handle:
        json.dump(payload, handle, indent=2)


def clear_progress(progress_path: Path) -> None:
    if progress_path.exists():
        progress_path.unlink()


def get_google_range_values(
    service,
    spreadsheet_id: str,
    range_name: str,
    value_render_option: str = "UNFORMATTED_VALUE",
) -> list[list[Any]]:
    response = (
        service.spreadsheets()
        .values()
        .get(
            spreadsheetId=spreadsheet_id,
            range=range_name,
            valueRenderOption=value_render_option,
            dateTimeRenderOption="FORMATTED_STRING",
        )
        .execute()
    )
    return response.get("values", [])


def get_google_single_value(
    service,
    spreadsheet_id: str,
    range_name: str,
    value_render_option: str = "UNFORMATTED_VALUE",
) -> Any:
    values = get_google_range_values(
        service,
        spreadsheet_id,
        range_name,
        value_render_option=value_render_option,
    )
    if values and values[0]:
        return values[0][0]
    return ""


def update_google_range_values(
    service,
    spreadsheet_id: str,
    range_name: str,
    values: list[list[str | int | float]],
    value_input_option: str = "RAW",
) -> None:
    (
        service.spreadsheets()
        .values()
        .update(
            spreadsheetId=spreadsheet_id,
            range=range_name,
            valueInputOption=value_input_option,
            body={"values": values},
        )
        .execute()
    )


def wait_for_numeric_dcf_output(
    service,
    spreadsheet_id: str,
    range_name: str,
    poll_seconds: float,
    timeout_seconds: float,
    min_wait_seconds: float,
    stable_reads_required: int,
) -> tuple[float | None, Any]:
    """Poll a DCF output cell until it returns the same finite number repeatedly.

    The minimum wait matters because SheetsFinance can briefly show a previous ticker's
    value, a loading string, or a formula error while the DCF tab recalculates.
    """
    start_time = time.time()
    last_number: float | None = None
    stable_reads = 0
    latest_raw: Any = ""
    stable_reads_required = max(1, stable_reads_required)
    poll_seconds = max(0.5, poll_seconds)

    while time.time() - start_time < timeout_seconds:
        elapsed = time.time() - start_time

        try:
            raw_output = get_google_single_value(
                service,
                spreadsheet_id,
                range_name,
                value_render_option="UNFORMATTED_VALUE",
            )
        except Exception as exc:
            raw_output = f"Google read error: {exc}"

        latest_raw = raw_output
        number = coerce_valid_number(raw_output)

        if elapsed >= min_wait_seconds and number is not None:
            if last_number is not None and abs(number - last_number) < 1e-12:
                stable_reads += 1
            else:
                last_number = number
                stable_reads = 1

            if stable_reads >= stable_reads_required:
                return number, raw_output
        elif number is None:
            last_number = None
            stable_reads = 0

        time.sleep(poll_seconds)

    return None, latest_raw


def refresh_valuation_gap_from_dcf(
    service,
    spreadsheet_id: str,
    tickers_sheet_name: str,
    dcf_sheet_name: str,
    valuation_gap_header: str,
    dcf_input_cell: str,
    dcf_output_cell: str,
    poll_seconds: float,
    timeout_seconds: float,
    min_wait_seconds: float,
    stable_reads_required: int,
    progress_path: Path,
    resume_progress: bool,
) -> tuple[int, list[str]]:
    source_values = get_google_values(service, spreadsheet_id, tickers_sheet_name)
    if not source_values:
        raise ValueError(f"Worksheet {tickers_sheet_name!r} does not contain data.")

    source_headers = [str(value).strip() for value in source_values[0]]
    ticker_idx = find_header_index(source_headers, TICKER_HEADER, tickers_sheet_name)
    valuation_gap_idx = find_header_index(source_headers, valuation_gap_header, tickers_sheet_name)
    valuation_gap_column = column_letter(valuation_gap_idx + 1)

    resume_row = 2
    if resume_progress:
        progress = load_progress(progress_path)
        if progress.get("source_worksheet") == tickers_sheet_name:
            try:
                stored_row = int(progress.get("last_completed_row", 1) or 1)
            except (TypeError, ValueError):
                stored_row = 1
            resume_row = max(2, stored_row + 1)
    else:
        clear_progress(progress_path)

    failures: list[str] = []
    refreshed_count = 0

    for row_number, row in enumerate(source_values[1:], start=2):
        if row_number < resume_row:
            continue

        padded_row = list(row) + [""] * max(0, len(source_headers) - len(row))
        ticker = str(padded_row[ticker_idx]).strip()
        if not ticker:
            break

        update_google_range_values(
            service,
            spreadsheet_id,
            f"{dcf_sheet_name}!{dcf_input_cell}",
            [[ticker]],
        )

        number, raw_output = wait_for_numeric_dcf_output(
            service=service,
            spreadsheet_id=spreadsheet_id,
            range_name=f"{dcf_sheet_name}!{dcf_output_cell}",
            poll_seconds=poll_seconds,
            timeout_seconds=timeout_seconds,
            min_wait_seconds=min_wait_seconds,
            stable_reads_required=stable_reads_required,
        )

        if number is None:
            failures.append(f"{ticker}: {short_sheet_value(raw_output)}")
            output_value: str | float = ""
        else:
            output_value = number

        # Always write something back. If DCF failed, write blank so old #DIV/0!,
        # #VALUE!, #ERROR!, or "Laster inn ..." values do not survive in the table.
        update_google_range_values(
            service,
            spreadsheet_id,
            f"{tickers_sheet_name}!{valuation_gap_column}{row_number}",
            [[output_value]],
        )

        refreshed_count += 1
        save_progress(progress_path, tickers_sheet_name, row_number, ticker)

    # If the script reached the end, do not keep a stale checkpoint that can make
    # the next clean run skip rows.
    clear_progress(progress_path)
    return refreshed_count, failures


def sanitize_valuation_gap_output(value: Any) -> str | float:
    number = coerce_valid_number(value)
    return number if number is not None else ""


def copy_google_results_to_output(
    service,
    spreadsheet_id: str,
    source_sheet_name: str,
    output_sheet_name: str,
    protected_output_columns: int,
    max_output_columns: int,
    valuation_gap_header: str = DEFAULT_VALUATION_GAP_HEADER,
) -> tuple[int, list[str], list[str]]:
    source_values = get_google_values(service, spreadsheet_id, source_sheet_name)
    output_values = get_google_values(service, spreadsheet_id, output_sheet_name)

    if not source_values:
        raise ValueError(f"Worksheet {source_sheet_name!r} does not contain data.")
    if not output_values:
        raise ValueError(f"Worksheet {output_sheet_name!r} does not contain data.")

    source_headers = [str(value).strip() for value in source_values[0]]
    output_headers = [str(value).strip() for value in output_values[0]]

    find_header_index(source_headers, TICKER_HEADER, source_sheet_name)

    mapped_headers: list[str] = []
    skipped_headers: list[str] = []

    protected_width = min(max(protected_output_columns, 0), len(output_headers), max_output_columns)
    source_width = min(len(source_headers), max_output_columns)
    total_width = max(protected_width, source_width)

    normalized_valuation_gap_header = normalize_header_name(valuation_gap_header)

    outgoing_values: list[list[str | int | float]] = []
    header_row: list[str | int | float] = [""] * total_width

    for idx in range(protected_width):
        header_row[idx] = output_headers[idx]

    for idx, header in enumerate(source_headers[protected_width:source_width], start=protected_width):
        header_row[idx] = header
        mapped_headers.append(header)

    outgoing_values.append(header_row)

    last_row = max(len(source_values), len(output_values))
    for row_idx in range(1, last_row):
        source_row = source_values[row_idx] if row_idx < len(source_values) else []
        output_row = output_values[row_idx] if row_idx < len(output_values) else []
        source_padded = list(source_row) + [""] * max(0, source_width - len(source_row))
        output_padded = list(output_row) + [""] * max(0, protected_width - len(output_row))

        merged_row: list[str | int | float] = [""] * total_width

        for idx in range(protected_width):
            merged_row[idx] = output_padded[idx]

        for idx in range(protected_width, source_width):
            value = source_padded[idx]
            header = source_headers[idx] if idx < len(source_headers) else ""
            if normalize_header_name(header) == normalized_valuation_gap_header:
                merged_row[idx] = sanitize_valuation_gap_output(value)
            else:
                merged_row[idx] = value

        outgoing_values.append(merged_row)

    last_column = column_letter(total_width)
    last_row_number = len(outgoing_values)
    (
        service.spreadsheets()
        .values()
        .update(
            spreadsheetId=spreadsheet_id,
            range=f"{output_sheet_name}!A1:{last_column}{last_row_number}",
            valueInputOption="RAW",
            body={"values": outgoing_values},
        )
        .execute()
    )

    return max(len(source_values), 1) - 1, mapped_headers, skipped_headers


def export_google_sheet_to_excel(
    service,
    spreadsheet_id: str,
    output_sheet_name: str,
    export_path: Path,
) -> tuple[int, int]:
    values = normalize_rows(get_google_values(service, spreadsheet_id, output_sheet_name))
    if not values:
        raise ValueError(f"Worksheet {output_sheet_name!r} does not contain data.")

    export_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    ws = workbook.active
    ws.title = DEFAULT_WORKSHEET_NAME

    for row in values:
        ws.append(row)

    workbook.save(export_path)
    return len(values), len(values[0]) if values else 0


def populate_google_sheet(service, spreadsheet_id: str, sheet_name: str) -> tuple[int, list[str], list[str]]:
    header_cells = get_google_header_cells(service, spreadsheet_id, sheet_name)
    if TICKER_HEADER not in header_cells:
        raise ValueError(f"Required header {TICKER_HEADER!r} not found in row 1.")

    missing_supported_headers = [header for header in FORMULA_BUILDERS if header not in header_cells]
    unmapped_headers = [header for header in header_cells if header not in FORMULA_BUILDERS and header != TICKER_HEADER]
    last_row = get_google_last_row(service, spreadsheet_id, sheet_name, header_cells[TICKER_HEADER])
    if last_row < 2:
        return 0, missing_supported_headers, unmapped_headers

    data = []
    for header, builder in FORMULA_BUILDERS.items():
        column = header_cells.get(header)
        if not column:
            continue

        values = []
        for row in range(2, last_row + 1):
            formula = builder(row, header_cells, last_row)
            values.append([formula if formula else ""])

        data.append(
            {
                "range": f"{sheet_name}!{column}2:{column}{last_row}",
                "values": values,
            }
        )

    if data:
        (
            service.spreadsheets()
            .values()
            .batchUpdate(
                spreadsheetId=spreadsheet_id,
                body={
                    "valueInputOption": "USER_ENTERED",
                    "data": data,
                },
            )
            .execute()
        )

    return last_row - 1, missing_supported_headers, unmapped_headers


def autosize_google_columns(service, spreadsheet_id: str, sheet_name: str, header_count: int) -> None:
    sheet_metadata = get_sheet_metadata(service, spreadsheet_id)
    sheet_id = get_google_sheet_id(sheet_metadata, sheet_name)
    (
        service.spreadsheets()
        .batchUpdate(
            spreadsheetId=spreadsheet_id,
            body={
                "requests": [
                    {
                        "autoResizeDimensions": {
                            "dimensions": {
                                "sheetId": sheet_id,
                                "dimension": "COLUMNS",
                                "startIndex": 0,
                                "endIndex": header_count,
                            }
                        }
                    }
                ]
            },
        )
        .execute()
    )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Google Sheets stock-ingestion workflow helpers.")
    parser.add_argument(
        "--mode",
        choices=("excel_formulas", "copy_to_output", "export_output", "sync_and_export"),
        default=os.getenv("INGESTION_MODE", "sync_and_export"),
        help="Workflow mode. Defaults to sync_and_export.",
    )
    parser.add_argument(
        "--sheet-id",
        default=os.getenv("GOOGLE_SHEET_ID", DEFAULT_SPREADSHEET_ID),
        help="Google Sheet ID to use for Google Sheets modes.",
    )
    parser.add_argument(
        "--credentials",
        default=str(GOOGLE_CREDENTIALS_PATH),
        help="Path to the Google service-account JSON file.",
    )
    parser.add_argument(
        "--worksheet",
        default=os.getenv("INGESTION_WORKSHEET", DEFAULT_WORKSHEET_NAME),
        help="Worksheet/tab name to update.",
    )
    parser.add_argument(
        "--source-worksheet",
        default=os.getenv("INGESTION_SOURCE_WORKSHEET", DEFAULT_SOURCE_WORKSHEET),
        help="Source worksheet/tab name for Google Sheets copy mode.",
    )
    parser.add_argument(
        "--output-worksheet",
        default=os.getenv("INGESTION_OUTPUT_WORKSHEET", DEFAULT_OUTPUT_WORKSHEET),
        help="Output worksheet/tab name for Google Sheets copy mode.",
    )
    parser.add_argument(
        "--protected-output-columns",
        type=int,
        default=int(os.getenv("INGESTION_PROTECTED_OUTPUT_COLUMNS", "0")),
        help="Number of leftmost output columns to leave untouched in Google Sheets copy mode.",
    )
    parser.add_argument(
        "--export-path",
        default=str(EXPORT_WORKBOOK_PATH),
        help="Excel export path for output-sheet export mode.",
    )
    parser.add_argument(
        "--max-output-columns",
        type=int,
        default=int(os.getenv("INGESTION_MAX_OUTPUT_COLUMNS", str(DEFAULT_MAX_OUTPUT_COLUMNS))),
        help="Maximum number of columns to copy into the output worksheet.",
    )
    parser.add_argument(
        "--dcf-worksheet",
        default=os.getenv("INGESTION_DCF_WORKSHEET", DEFAULT_DCF_WORKSHEET),
        help="DCF worksheet/tab name used to calculate valuation_gap.",
    )
    parser.add_argument(
        "--valuation-gap-header",
        default=os.getenv("INGESTION_VALUATION_GAP_HEADER", DEFAULT_VALUATION_GAP_HEADER),
        help="Header name for valuation_gap in the source worksheet.",
    )
    parser.add_argument(
        "--dcf-input-cell",
        default=os.getenv("INGESTION_DCF_INPUT_CELL", DEFAULT_DCF_INPUT_CELL),
        help="DCF input cell that receives the ticker.",
    )
    parser.add_argument(
        "--dcf-output-cell",
        default=os.getenv("INGESTION_DCF_OUTPUT_CELL", DEFAULT_DCF_OUTPUT_CELL),
        help="DCF output cell that returns valuation_gap.",
    )
    parser.add_argument(
        "--dcf-poll-seconds",
        type=float,
        default=float(os.getenv("INGESTION_DCF_POLL_SECONDS", str(DEFAULT_DCF_POLL_SECONDS))),
        help="Seconds between DCF polling attempts.",
    )
    parser.add_argument(
        "--dcf-timeout-seconds",
        type=float,
        default=float(os.getenv("INGESTION_DCF_TIMEOUT_SECONDS", str(DEFAULT_DCF_TIMEOUT_SECONDS))),
        help="Maximum seconds to wait per ticker for DCF recalculation.",
    )
    parser.add_argument(
        "--dcf-min-wait-seconds",
        type=float,
        default=float(os.getenv("INGESTION_DCF_MIN_WAIT_SECONDS", str(DEFAULT_DCF_MIN_WAIT_SECONDS))),
        help="Minimum seconds to wait before accepting a DCF output value.",
    )
    parser.add_argument(
        "--dcf-stable-reads",
        type=int,
        default=int(os.getenv("INGESTION_DCF_STABLE_READS", str(DEFAULT_DCF_STABLE_READS))),
        help="Number of identical numeric reads required before accepting DCF output.",
    )
    parser.add_argument(
        "--progress-path",
        default=str(PROGRESS_PATH),
        help="Local checkpoint file used to resume valuation_gap processing.",
    )
    parser.add_argument(
        "--no-resume-progress",
        action="store_true",
        help="Ignore the checkpoint file and start valuation_gap refresh from row 2.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    sheet_name = args.worksheet

    if args.mode == "excel_formulas":
        workbook = load_workbook(WORKBOOK_PATH)
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Worksheet {sheet_name!r} not found in {WORKBOOK_PATH}")

        ws = workbook[sheet_name]
        updated_rows, missing_supported_headers, unmapped_headers = populate_formulas(ws)
        autosize_columns(ws)
        workbook.save(WORKBOOK_PATH)

        print(f"Updated workbook: {WORKBOOK_PATH}")
        print(f"Worksheet: {sheet_name}")
        print(f"Rows with ticker formulas refreshed: {updated_rows}")
        print("Open the workbook in Excel with the SheetsFinance add-in enabled to calculate the formulas.")

        if missing_supported_headers:
            print("Known columns not present in this workbook:")
            for header in missing_supported_headers:
                print(f"- {header}")

        if unmapped_headers:
            print("Headers left untouched because no formula mapping is defined:")
            for header in unmapped_headers:
                print(f"- {header}")

        return

    if not args.sheet_id:
        raise ValueError("Missing Google Sheet ID. Pass --sheet-id or set GOOGLE_SHEET_ID.")

    service = load_google_service(Path(args.credentials))

    if args.mode in {"copy_to_output", "sync_and_export"}:
        refreshed_rows, failed_tickers = refresh_valuation_gap_from_dcf(
            service=service,
            spreadsheet_id=args.sheet_id,
            tickers_sheet_name=args.source_worksheet,
            dcf_sheet_name=args.dcf_worksheet,
            valuation_gap_header=args.valuation_gap_header,
            dcf_input_cell=args.dcf_input_cell,
            dcf_output_cell=args.dcf_output_cell,
            poll_seconds=args.dcf_poll_seconds,
            timeout_seconds=args.dcf_timeout_seconds,
            min_wait_seconds=args.dcf_min_wait_seconds,
            stable_reads_required=args.dcf_stable_reads,
            progress_path=Path(args.progress_path),
            resume_progress=not args.no_resume_progress,
        )

        updated_rows, copied_headers, skipped_headers = copy_google_results_to_output(
            service=service,
            spreadsheet_id=args.sheet_id,
            source_sheet_name=args.source_worksheet,
            output_sheet_name=args.output_worksheet,
            protected_output_columns=args.protected_output_columns,
            max_output_columns=args.max_output_columns,
            valuation_gap_header=args.valuation_gap_header,
        )

        autosize_google_columns(service, args.sheet_id, args.output_worksheet, args.max_output_columns)

        print(f"Updated Google Sheet ID: {args.sheet_id}")
        print(f"Source worksheet: {args.source_worksheet}")
        print(f"Output worksheet: {args.output_worksheet}")
        print(f"DCF worksheet: {args.dcf_worksheet}")
        print(f"Valuation gap rows refreshed in source: {refreshed_rows}")
        print(f"Rows copied: {updated_rows}")
        if args.protected_output_columns > 0:
            print(f"Protected output columns left untouched: A:{column_letter(args.protected_output_columns)}")
        else:
            print("Protected output columns left untouched: none")
        print(f"Maximum copied column: {column_letter(args.max_output_columns)}")
        print(f"DCF timeout per ticker: {args.dcf_timeout_seconds} seconds")
        print(f"DCF minimum wait per ticker: {args.dcf_min_wait_seconds} seconds")
        print(f"DCF stable numeric reads required: {args.dcf_stable_reads}")

        if copied_headers:
            print("Copied output headers:")
            for header in copied_headers:
                print(f"- {header}")

        if skipped_headers:
            print("Output headers skipped because they were not found in the source worksheet:")
            for header in skipped_headers:
                print(f"- {header}")

        if failed_tickers:
            print("Tickers where DCF valuation_gap was not a valid number and was written as blank:")
            for ticker in failed_tickers:
                print(f"- {ticker}")

        if args.mode == "sync_and_export":
            row_count, column_count = export_google_sheet_to_excel(
                service,
                args.sheet_id,
                args.output_worksheet,
                Path(args.export_path),
            )
            print(f"Exported worksheet: {args.output_worksheet}")
            print(f"Export path: {args.export_path}")
            print(f"Rows exported: {row_count}")
            print(f"Columns exported: {column_count}")

        return

    if args.mode == "export_output":
        row_count, column_count = export_google_sheet_to_excel(
            service,
            args.sheet_id,
            args.output_worksheet,
            Path(args.export_path),
        )
        print(f"Exported worksheet: {args.output_worksheet}")
        print(f"Export path: {args.export_path}")
        print(f"Rows exported: {row_count}")
        print(f"Columns exported: {column_count}")
        return

    raise ValueError(f"Unsupported mode: {args.mode}")


if __name__ == "__main__":
    main()
